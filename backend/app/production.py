from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Any

from influxdb_client import Point, WritePrecision
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from .influx import parse_timestamp


REQUIRED_FIELDS = ("factory", "process", "quantity")

HEADER_ALIASES = {
    "factory": "factory",
    "site": "factory",
    "plant": "factory",
    "process": "process",
    "line": "process",
    "product": "product",
    "item": "product",
    "shift": "shift",
    "quantity": "quantity",
    "qty": "quantity",
    "production": "quantity",
    "time": "time",
    "timestamp": "time",
    "date": "time",
    "datetime": "time",
    "note": "note",
}


@dataclass(frozen=True)
class ProductionRecord:
    factory: str
    process: str
    quantity: float
    time: datetime
    product: str | None = None
    shift: str | None = None
    note: str | None = None


def _clean_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_quantity(value: Any) -> float:
    if value in (None, ""):
        raise ValueError("quantity is required")
    try:
        quantity = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("quantity must be numeric") from exc
    if quantity < 0:
        raise ValueError("quantity must be zero or greater")
    return quantity


def parse_production_payload(payload: dict[str, Any]) -> ProductionRecord:
    missing = [field for field in REQUIRED_FIELDS if payload.get(field) in (None, "")]
    if missing:
        raise ValueError(f"missing required fields: {', '.join(missing)}")

    return ProductionRecord(
        factory=str(payload["factory"]).strip(),
        process=str(payload["process"]).strip(),
        quantity=_parse_quantity(payload["quantity"]),
        time=parse_timestamp(payload.get("time")),
        product=_clean_string(payload.get("product")),
        shift=_clean_string(payload.get("shift")),
        note=_clean_string(payload.get("note")),
    )


def production_record_to_point(record: ProductionRecord) -> Point:
    point = (
        Point("production_input")
        .tag("factory", record.factory)
        .tag("process", record.process)
        .field("quantity", record.quantity)
        .time(record.time, WritePrecision.NS)
    )

    if record.product:
        point.tag("product", record.product)
    if record.shift:
        point.tag("shift", record.shift)
    if record.note:
        point.field("note", record.note)

    return point


def _normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    key = str(value).strip().lower()
    return HEADER_ALIASES.get(key, key)


def parse_production_excel(file: FileStorage) -> tuple[list[ProductionRecord], list[dict[str, Any]]]:
    workbook = load_workbook(file.stream, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("Excel file is empty") from None

    headers = [_normalize_header(value) for value in header_row]
    header_set = {header for header in headers if header}
    missing = [field for field in REQUIRED_FIELDS if field not in header_set]
    if missing:
        raise ValueError(f"missing required columns: {', '.join(missing)}")

    records: list[ProductionRecord] = []
    errors: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row):
            continue

        payload = {
            header: row[position]
            for position, header in enumerate(headers)
            if header and position < len(row)
        }

        try:
            records.append(parse_production_payload(payload))
        except ValueError as exc:
            errors.append({"row": index, "error": str(exc)})

    return records, errors
